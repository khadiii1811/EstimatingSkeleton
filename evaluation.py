import statistics
import datetime
import psutil
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, Toplevel, messagebox, StringVar, Text, Scrollbar, Label, Frame, Button, WORD, DISABLED, END, RIGHT, Y, BOTH, X, LEFT

# Global variables for performance tracking and evaluation
evaluation_data = {
    "timestamp": "",
    "features_used": [],
    "performance": {
        "fps_values": [],
        "avg_fps": 0,
        "min_fps": 0,
        "max_fps": 0,
        "std_fps": 0,
        "stability_score": 0  # Stability score (low=unstable, high=stable)
    },
    "accuracy": {
        "skeleton_detection_rate": 0,
        "face_detection_rate": 0,
        "hand_detection_rate": 0,
        "fall_detection_rate": 0,
        "detection_confidence": 0  # Detection confidence
    },
    "resource_usage": {
        "cpu_usage": [],
        "memory_usage": [],
        "avg_cpu": 0,
        "avg_memory": 0,
        "efficiency_score": 0  # Resource efficiency score
    },
    "usability": {
        "response_time": 0,  # Average response time (ms)
        "usability_score": 0  # User friendliness score
    },
    "overall_score": 0,  # Composite score
    "notes": ""
}

# Global variable to store performance data for all sessions
session_history = []

# Function to reset evaluation data
def reset_evaluation_data():
    global evaluation_data
    evaluation_data = {
        "timestamp": "",
        "features_used": [],
        "performance": {
            "fps_values": [],
            "avg_fps": 0,
            "min_fps": 0,
            "max_fps": 0,
            "std_fps": 0,
            "stability_score": 0
        },
        "accuracy": {
            "skeleton_detection_rate": 0,
            "face_detection_rate": 0,
            "hand_detection_rate": 0,
            "fall_detection_rate": 0,
            "detection_confidence": 0
        },
        "resource_usage": {
            "cpu_usage": [],
            "memory_usage": [],
            "avg_cpu": 0,
            "avg_memory": 0,
            "efficiency_score": 0
        },
        "usability": {
            "response_time": 0,
            "usability_score": 0
        },
        "overall_score": 0,
        "notes": ""
    }

def show_evaluation_report(root, panel_color, text_color, accent_color, button_color_primary, button_color_danger, btn_font):
    """Display a detailed evaluation report in a separate window"""
    
    # Check if there's evaluation data
    if not session_history:
        messagebox.showinfo("Notice", "No evaluation data yet. Please use features before viewing report.")
        return
    
    # Create report window
    report_window = Toplevel(root)
    report_window.title("Performance Evaluation Report")
    report_window.geometry("900x700")
    report_window.configure(bg=panel_color)
    
    # Report frame
    report_frame = Frame(report_window, bg=panel_color, padx=20, pady=20)
    report_frame.pack(fill=BOTH, expand=True)
    
    # Title
    Label(report_frame, text="PERFORMANCE EVALUATION REPORT", 
          font=('Helvetica', 18, 'bold'), bg=panel_color, fg=accent_color).pack(pady=10)
    
    # Report generation time
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    Label(report_frame, text=f"Report generated: {current_time}", 
          font=('Helvetica', 12), bg=panel_color, fg=text_color).pack(pady=5)
    
    # Report content frame
    content_frame = Frame(report_frame, bg='#252530', padx=15, pady=15)
    content_frame.pack(fill=BOTH, expand=True, pady=10)
    
    # Create Text widget with scrollbar
    text_scroll = Scrollbar(content_frame)
    text_scroll.pack(side=RIGHT, fill=Y)
    
    report_text = Text(content_frame, wrap=WORD, bg='#252530', fg=text_color,
                      font=('Consolas', 11), padx=10, pady=10,
                      yscrollcommand=text_scroll.set)
    report_text.pack(fill=BOTH, expand=True)
    text_scroll.config(command=report_text.yview)
    
    # Insert report content
    report_text.insert(END, "===== EVALUATION CRITERIA =====\n\n")
    
    # Add evaluation table and criteria
    report_text.insert(END, "SCORING STANDARD:\n")
    report_text.insert(END, "  • Overall Score: Scale 0-10\n")
    report_text.insert(END, "    - 9-10: Excellent\n")
    report_text.insert(END, "    - 7-8.9: Good\n")
    report_text.insert(END, "    - 5-6.9: Fair\n")
    report_text.insert(END, "    - 3-4.9: Average\n")
    report_text.insert(END, "    - 0-2.9: Poor\n\n")
    
    report_text.insert(END, "MAIN EVALUATION CRITERIA:\n")
    report_text.insert(END, "1. Performance (40%)\n")
    report_text.insert(END, "   - Average FPS: Frame processing speed (higher = better)\n")
    report_text.insert(END, "   - Stability Score: FPS stability (1-10)\n\n")
    
    report_text.insert(END, "2. Accuracy (30%)\n")
    report_text.insert(END, "   - Detection Rate: Ratio of successful detection frames\n")
    report_text.insert(END, "   - Detection Confidence: Confidence level (1-10)\n\n")
    
    report_text.insert(END, "3. Resource Efficiency (20%)\n")
    report_text.insert(END, "   - Average CPU: CPU usage (%)\n")
    report_text.insert(END, "   - Average Memory: RAM usage (MB)\n")
    report_text.insert(END, "   - Efficiency Score: Resource usage efficiency (1-10)\n\n")
    
    report_text.insert(END, "4. Usability (10%)\n")
    report_text.insert(END, "   - Response Time: Average processing time (ms)\n")
    report_text.insert(END, "   - Usability Score: User-friendliness (1-10)\n\n")
    
    report_text.insert(END, "===== OVERVIEW =====\n\n")
    report_text.insert(END, f"Total sessions: {len(session_history)}\n")
    
    # Calculate overall metrics
    all_fps = []
    features_used = set()
    overall_scores = []
    for session in session_history:
        all_fps.extend(session["performance"]["fps_values"])
        features_used.update(session["features_used"])
        overall_scores.append(session["overall_score"])
    
    avg_overall_score = statistics.mean(overall_scores) if overall_scores else 0
    
    # Classify overall performance
    performance_category = ""
    if avg_overall_score >= 9:
        performance_category = "Excellent"
    elif avg_overall_score >= 7:
        performance_category = "Good"
    elif avg_overall_score >= 5:
        performance_category = "Fair"
    elif avg_overall_score >= 3:
        performance_category = "Average"
    else:
        performance_category = "Poor"
    
    report_text.insert(END, f"Features used: {', '.join(features_used)}\n")
    report_text.insert(END, f"Overall average FPS: {statistics.mean(all_fps):.2f} fps\n")
    report_text.insert(END, f"Overall evaluation score: {avg_overall_score:.2f}/10 ({performance_category})\n\n")
    
    # Session details
    report_text.insert(END, "===== SESSION DETAILS =====\n\n")
    
    for i, session in enumerate(session_history):
        report_text.insert(END, f"Session {i+1} - {session['timestamp']}\n")
        report_text.insert(END, f"Features used: {', '.join(session['features_used'])}\n")
        
        # 1. Performance information
        report_text.insert(END, "1. Performance:\n")
        report_text.insert(END, f"   - Average FPS: {session['performance']['avg_fps']:.2f} fps\n")
        report_text.insert(END, f"   - Minimum FPS: {session['performance']['min_fps']:.2f} fps\n")
        report_text.insert(END, f"   - Maximum FPS: {session['performance']['max_fps']:.2f} fps\n")
        report_text.insert(END, f"   - FPS standard deviation: {session['performance']['std_fps']:.2f}\n")
        report_text.insert(END, f"   - Stability score: {session['performance']['stability_score']:.2f}/10\n")
        
        # 2. Accuracy information
        report_text.insert(END, "2. Accuracy:\n")
        if 'pose_estimation' in session['features_used'] or 'fall_detection' in session['features_used']:
            report_text.insert(END, f"   - Skeleton detection rate: {session['accuracy']['skeleton_detection_rate']:.2%}\n")
        if 'face_detection' in session['features_used']:
            report_text.insert(END, f"   - Face detection rate: {session['accuracy']['face_detection_rate']:.2%}\n")
        if 'hand_tracking' in session['features_used']:
            report_text.insert(END, f"   - Hand detection rate: {session['accuracy']['hand_detection_rate']:.2%}\n")
        report_text.insert(END, f"   - Detection confidence: {session['accuracy']['detection_confidence']:.2f}/10\n")
        
        # 3. Resource efficiency
        report_text.insert(END, "3. Resource Efficiency:\n")
        report_text.insert(END, f"   - Average CPU: {session['resource_usage']['avg_cpu']:.2f}%\n")
        report_text.insert(END, f"   - Average memory: {session['resource_usage']['avg_memory']:.2f} MB\n")
        report_text.insert(END, f"   - Resource efficiency score: {session['resource_usage']['efficiency_score']:.2f}/10\n")
        
        # 4. Usability
        report_text.insert(END, "4. Usability:\n")
        report_text.insert(END, f"   - Average response time: {session['usability']['response_time']:.2f} ms\n")
        report_text.insert(END, f"   - User-friendliness score: {session['usability']['usability_score']:.2f}/10\n")
        
        # Overall score
        performance_category = ""
        if session['overall_score'] >= 9:
            performance_category = "Excellent"
        elif session['overall_score'] >= 7:
            performance_category = "Good"
        elif session['overall_score'] >= 5:
            performance_category = "Fair"
        elif session['overall_score'] >= 3:
            performance_category = "Average"
        else:
            performance_category = "Poor"
        
        report_text.insert(END, f"\nOverall score: {session['overall_score']:.2f}/10 ({performance_category})\n\n")
        report_text.insert(END, "--------------------------------------------------------\n\n")
    
    # Add notes or assessments
    report_text.insert(END, "===== EVALUATION AND NOTES =====\n\n")
    
    # Feature-specific analysis
    report_text.insert(END, "Detailed analysis by feature:\n\n")
    
    # (The rest of this function contains the detailed analysis by feature, 
    # feature comparison, recommendations, etc. For brevity, I'll omit that part)
    
    # Make text readonly
    report_text.config(state=DISABLED)
    
    # Frame for buttons
    button_frame = Frame(report_frame, bg=panel_color, pady=10)
    button_frame.pack(fill=X)
    
    # PDF export function
    def export_report_to_pdf():
        export_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            title="Save report as PDF"
        )
        
        if not export_path:
            return
        
        try:
            # Create PDF document
            doc = SimpleDocTemplate(export_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Add report title
            title_style = styles['Title']
            story.append(Paragraph("PERFORMANCE EVALUATION REPORT", title_style))
            story.append(Spacer(1, 0.25*inch))
            
            # Add report generation info
            normal_style = styles['Normal']
            story.append(Paragraph(f"Report generated: {current_time}", normal_style))
            story.append(Spacer(1, 0.25*inch))
            
            # Add evaluation standards
            heading_style = styles['Heading1']
            story.append(Paragraph("EVALUATION STANDARDS", heading_style))
            
            # Standards table
            data = [
                ["Score Range", "Classification"],
                ["9.0 - 10.0", "Excellent"],
                ["7.0 - 8.9", "Good"],
                ["5.0 - 6.9", "Fair"],
                ["3.0 - 4.9", "Average"],
                ["0.0 - 2.9", "Poor"]
            ]
            
            table = Table(data, colWidths=[1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, 1), colors.lightgreen),
                ('BACKGROUND', (0, 2), (-1, 2), colors.lightgreen),
                ('BACKGROUND', (0, 3), (-1, 3), colors.lightyellow),
                ('BACKGROUND', (0, 4), (-1, 4), colors.lightyellow),
                ('BACKGROUND', (0, 5), (-1, 5), colors.pink),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 0.25*inch))
            
            # Add the main criteria
            story.append(Paragraph("MAIN EVALUATION CRITERIA", heading_style))
            
            # (Rest of PDF export logic...)
            
            # Build the PDF
            doc.build(story)
            messagebox.showinfo("Notice", f"PDF report exported successfully to:\n{export_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not export PDF report: {str(e)}")
    
    # Excel export function
    def export_report_to_excel():
        export_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save report as Excel"
        )
        
        if not export_path:
            return
        
        try:
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Create sheets
            overview_sheet = wb.active
            overview_sheet.title = "Overview"
            session_sheet = wb.create_sheet("Session Details")
            feature_sheet = wb.create_sheet("Feature Analysis")
            chart_data_sheet = wb.create_sheet("Chart Data")
            
            # Setup styles
            title_font = Font(name='Calibri', size=14, bold=True)
            header_font = Font(name='Calibri', size=12, bold=True)
            normal_font = Font(name='Calibri', size=11)
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Fill colors
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            alt_row_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            # (Rest of Excel export logic...)
            
            # Save the Excel file
            wb.save(export_path)
            messagebox.showinfo("Notice", f"Excel report exported successfully to:\n{export_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not export Excel report: {str(e)}")
    
    # Add export buttons
    Button(button_frame, text="Export to PDF", font=btn_font,
          bg=button_color_primary, fg=text_color, padx=10, pady=5,
          command=export_report_to_pdf).pack(side=LEFT, padx=5)
    
    Button(button_frame, text="Export to Excel", font=btn_font,
          bg='#4CAF50', fg='white', padx=10, pady=5,
          command=export_report_to_excel).pack(side=LEFT, padx=5)
    
    Button(button_frame, text="Close", font=btn_font,
          bg=button_color_danger, fg=text_color, padx=10, pady=5,
          command=report_window.destroy).pack(side=RIGHT, padx=5)
    
    # Make report window top-level
    report_window.transient(root)
    report_window.grab_set()
    
    # Center the report window
    report_window.update_idletasks()
    width = report_window.winfo_width()
    height = report_window.winfo_height()
    x = (report_window.winfo_screenwidth() // 2) - (width // 2)
    y = (report_window.winfo_screenheight() // 2) - (height // 2)
    report_window.geometry('{}x{}+{}+{}'.format(width, height, x, y))

def show_performance_comparison(root, panel_color, text_color, accent_color, button_color_primary, button_color_danger, btn_font):
    """Display a comparison of performance between features"""
    # Implementation of performance comparison visualization
    # Similar structure to show_evaluation_report but with chart visualization
    pass 